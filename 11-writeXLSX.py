# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from os.path import *
from datetime import date, datetime
from models import db, objects

dbConnection = db.Database()
breaksDB = dbConnection.connect()
gamesObj = objects.Games(breaksDB)
tournamentsObj = objects.Tournaments(breaksDB)
playersObj = objects.Players(breaksDB)
dayDateTime = date.today()
dayString = dayDateTime.strftime("%Y-%m-%d")
filepath = "xlsx/{}.xlsx".format(dayString)

if isfile(filepath):
    workbook = openpyxl.load_workbook(filepath)
else:
    workbook = openpyxl.Workbook()

fontRegularBlack = Font(
    name = "Arial",
    size = 10,
    bold = False
)
fontBoldBlack = Font(
    name = "Arial",
    size = 10,
    bold = True
)
fontRegularWhite = Font(
    name = "Arial",
    size = 10,
    bold = False,
    color = "ffffff"
)
alignmentLeft = Alignment(
    horizontal = "left",
    vertical = "center"
)
alignmentCenter = Alignment(
    horizontal = "center",
    vertical = "center"
)
rightBorder = Border(
    right = Side(style="thin")
)
categories = {
    'ATP-250': {'backColor': "1eaeea", 'foreFont': fontRegularWhite},
    'ATP-500': {'backColor': "ccccdd", 'foreFont': fontRegularBlack},
    'ATP-1000': {'backColor': "bb9640", 'foreFont': fontRegularWhite},
    'ATP-GS': {'backColor': "20305a", 'foreFont': fontRegularWhite},
    'CH': {'backColor': "53a01d", 'foreFont': fontRegularWhite},
    'ITF': {'backColor': "f8c408", 'foreFont': fontRegularBlack},
    'WTA': {'backColor': "7814ff", 'foreFont': fontRegularWhite}
}
categoriesToCombine = ["CH", "ITF"]
surfaces = {'D': "788575", 'T': "ff7500", 'H': "74ae30", 'I': "618cb1"}

worksheet = workbook.active
worksheet.title = "Daily games"
worksheet['A1'] = dayDateTime
worksheet['A1'].number_format = "dd-MMM"
worksheet['A1'].font = fontBoldBlack
worksheet['A1'].alignment = alignmentCenter
worksheet['A1'].border = rightBorder
worksheet['A1'].fill = PatternFill(start_color = "ffd966", end_color = "ffd966", fill_type = "solid")
games = gamesObj.find_all([{'gameDay': dayString, 'odd': {"$exists": True, "$gte": 1.57}}])

for gameIndex, game in enumerate(games):
    numRow = gameIndex + 1
    tournament = tournamentsObj.find([{'_id': game['tournament']}])
    player1 = playersObj.find([{'_id': game['player1ID']}])
    player2 = playersObj.find([{'_id': game['player2ID']}])

    if tournament['category'] == "ATP":
        categoryKey = "ATP-{}".format(tournament['subcategory'])
    else:
        categoryKey = tournament['category']

    worksheet['B{}'.format(numRow)] = tournament['category']
    worksheet['B{}'.format(numRow)].font = categories[categoryKey]['foreFont']
    worksheet['B{}'.format(numRow)].alignment = alignmentCenter
    worksheet['B{}'.format(numRow)].fill = PatternFill(start_color = categories[categoryKey]['backColor'], end_color = categories[categoryKey]['backColor'], fill_type = "solid")

    if tournament['category'] in categoriesToCombine:
        worksheet.merge_cells('B{}:C{}'.format(numRow, numRow))
    else:
        worksheet['C{}'.format(numRow)] = tournament['subcategory']
        worksheet['C{}'.format(numRow)].font = categories[categoryKey]['foreFont']
        worksheet['C{}'.format(numRow)].alignment = alignmentCenter
        worksheet['C{}'.format(numRow)].fill = PatternFill(start_color = categories[categoryKey]['backColor'], end_color = categories[categoryKey]['backColor'], fill_type = "solid")

    worksheet['D{}'.format(numRow)] = tournament['surface']
    worksheet['D{}'.format(numRow)].font = fontRegularWhite
    worksheet['D{}'.format(numRow)].alignment = alignmentCenter
    worksheet['D{}'.format(numRow)].fill = PatternFill(start_color = surfaces[tournament['surface']], end_color = surfaces[tournament['surface']], fill_type = "solid")
    worksheet['E{}'.format(numRow)] = " ".join(game['FS-player1'].split(" ")[0:-1])
    worksheet['E{}'.format(numRow)].font = fontBoldBlack
    worksheet['E{}'.format(numRow)].alignment = alignmentLeft
    worksheet['F{}'.format(numRow)] = " ".join(game['FS-player2'].split(" ")[0:-1])
    worksheet['F{}'.format(numRow)].font = fontRegularBlack
    worksheet['F{}'.format(numRow)].alignment = alignmentLeft
    worksheet['G{}'.format(numRow)] = "OK"
    worksheet['G{}'.format(numRow)].font = fontRegularWhite
    worksheet['G{}'.format(numRow)].alignment = alignmentCenter
    worksheet['G{}'.format(numRow)].fill = PatternFill(start_color = "34a853", end_color = "34a853", fill_type = "solid")
    worksheet['H{}'.format(numRow)] = 1.01
    worksheet['H{}'.format(numRow)].font = fontRegularBlack
    worksheet['H{}'.format(numRow)].alignment = alignmentCenter

    if player1['country'] == tournament['country']:
        player1['home'] = "S"
    else:
        player1['home'] = "N"

    if player2['country'] == tournament['country']:
        player2['home'] = "S"
    else:
        player2['home'] = "N"

    if player1['home'] == player2['home']:
        backColor = "ffffff"
        foreFont = fontRegularBlack
    elif player1['home'] == "S":
        backColor = "34a853"
        foreFont = fontRegularWhite
    else:
        backColor = "ee0000"
        foreFont = fontRegularWhite

    worksheet['I{}'.format(numRow)] = player1['home']
    worksheet['I{}'.format(numRow)].font = foreFont
    worksheet['I{}'.format(numRow)].alignment = alignmentCenter
    worksheet['J{}'.format(numRow)] = player2['home']
    worksheet['J{}'.format(numRow)].font = foreFont
    worksheet['J{}'.format(numRow)].alignment = alignmentCenter

    if backColor != "ffffff":
        worksheet['I{}'.format(numRow)].fill = PatternFill(start_color = backColor, end_color = backColor, fill_type = "solid")
        worksheet['J{}'.format(numRow)].fill = PatternFill(start_color = backColor, end_color = backColor, fill_type = "solid")

    worksheet['K{}'.format(numRow)] = game['odd']
    worksheet['K{}'.format(numRow)].font = fontRegularBlack
    worksheet['K{}'.format(numRow)].alignment = alignmentCenter
    worksheet['K{}'.format(numRow)].number_format = "#,##0.00"
    worksheet['L{}'.format(numRow)].value = str(game['probability']).replace(".", ",") + "%"
    worksheet['L{}'.format(numRow)].number_format = "0.00%"
    worksheet['L{}'.format(numRow)].font = fontBoldBlack
    worksheet['L{}'.format(numRow)].alignment = alignmentCenter
    worksheet['M{}'.format(numRow)].value = "=1/K{}".format(numRow) # 62,11%
    worksheet['M{}'.format(numRow)].number_format = "0.00%"
    worksheet['M{}'.format(numRow)].font = fontRegularBlack
    worksheet['M{}'.format(numRow)].alignment = alignmentCenter
    worksheet['N{}'.format(numRow)] = "=L{}/M{}".format(numRow, numRow)
    worksheet['N{}'.format(numRow)].font = fontBoldBlack
    worksheet['N{}'.format(numRow)].alignment = alignmentCenter
    worksheet['N{}'.format(numRow)].number_format = "#,##0.00"

worksheet.merge_cells('A1:A{}'.format(numRow))
workbook.save(filepath)

'''
    # TODO

    1) Into player lastGames (or game documents), newField: last ("OK", "01", "02", "10", "11", "12", "20", "21", "22")
    2) Opponent win odd
    3) Print avgProbability in 07-printDailyGames
    4) Refactoring 04-getLastGames calling tennisExplorer.getLastGamesByPlayer
'''
