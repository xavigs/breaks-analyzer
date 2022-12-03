# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from os.path import *
from datetime import date, datetime, timedelta
from models import db, objects

dbConnection = db.Database()
breaksDB = dbConnection.connect()
systemsObj = objects.Systems(breaksDB)
systems = systemsObj.find_all()
dayDateTime = date.today()
dayString = dayDateTime.strftime("%Y-%m-%d")
filepath = "systems{}.xlsx".format(dayString)

if isfile(filepath):
    workbook = openpyxl.load_workbook(filepath)
else:
    workbook = openpyxl.Workbook()

fontRegularBlack = Font(
    name = "Arial",
    size = 10,
    bold = False
)
fontRegularWhite = Font(
    name = "Arial",
    size = 10,
    bold = False,
    color = "ffffff"
)
fontBoldBlack = Font(
    name = "Arial",
    size = 10,
    bold = True
)
fontBoldWhite = Font(
    name = "Arial",
    size = 10,
    bold = True,
    color = "ffffff"
)
alignmentLeft = Alignment(
    horizontal = "left",
    vertical = "center"
)
alignmentCenter = Alignment(
    horizontal = "center",
    vertical = "center"
)

worksheet = workbook.active
worksheet.title = "Systems"
worksheet['A1'] = "System"
worksheet['A1'].font = fontBoldWhite
worksheet['A1'].alignment = alignmentLeft
worksheet['A1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['B1'] = "Start"
worksheet['B1'].font = fontBoldWhite
worksheet['B1'].alignment = alignmentCenter
worksheet['B1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['C1'] = "Months >10%"
worksheet['C1'].font = fontBoldWhite
worksheet['C1'].alignment = alignmentCenter
worksheet['C1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet.merge_cells('C1:D1')
worksheet['E1'] = "Positive months"
worksheet['E1'].font = fontBoldWhite
worksheet['E1'].alignment = alignmentCenter
worksheet['E1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet.merge_cells('E1:F1')
worksheet['G1'] = "Nº months"
worksheet['G1'].font = fontBoldWhite
worksheet['G1'].alignment = alignmentCenter
worksheet['G1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['H1'] = "Picks"
worksheet['H1'].font = fontBoldWhite
worksheet['H1'].alignment = alignmentCenter
worksheet['H1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['I1'] = "Picks/month"
worksheet['I1'].font = fontBoldWhite
worksheet['I1'].alignment = alignmentCenter
worksheet['I1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet.merge_cells('I1:J1')
worksheet['K1'] = "Yield TOT"
worksheet['K1'].font = fontBoldWhite
worksheet['K1'].alignment = alignmentCenter
worksheet['K1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet.merge_cells('K1:L1')
worksheet['M1'] = "Start bank"
worksheet['M1'].font = fontBoldWhite
worksheet['M1'].alignment = alignmentCenter
worksheet['M1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['N1'] = "End bank"
worksheet['N1'].font = fontBoldWhite
worksheet['N1'].alignment = alignmentCenter
worksheet['N1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['O1'] = "Profit"
worksheet['O1'].font = fontBoldWhite
worksheet['O1'].alignment = alignmentCenter
worksheet['O1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['P1'] = "Monthly\nProfit"
worksheet['P1'].font = fontBoldWhite
worksheet['P1'].alignment = alignmentCenter
worksheet['P1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['Q1'] = "Yearly\nProfit"
worksheet['Q1'].font = fontBoldWhite
worksheet['Q1'].alignment = alignmentCenter
worksheet['Q1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['R1'] = "EVAL"
worksheet['R1'].font = fontBoldWhite
worksheet['R1'].alignment = alignmentCenter
worksheet['R1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['T1'] = "GLOB"
worksheet['T1'].font = fontBoldWhite
worksheet['T1'].alignment = alignmentCenter
worksheet['T1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['U1'] = "AVG"
worksheet['U1'].font = fontBoldWhite
worksheet['U1'].alignment = alignmentCenter
worksheet['U1'].fill = PatternFill(start_color = "5b277d", end_color = "5b277d", fill_type = "solid")
worksheet['V1'] = "PREMIUN TIPSTERLAND"
worksheet['V1'].font = fontBoldWhite
worksheet['V1'].alignment = alignmentCenter
worksheet['V1'].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
worksheet.merge_cells('V1:Z1')
row = 2

for systemIndex, systemData in enumerate(systems):
    worksheet['A{}'.format(row)] = systemData['name']
    worksheet['A{}'.format(row)].font = fontRegularBlack
    worksheet['A{}'.format(row)].alignment = alignmentLeft
    worksheet['B{}'.format(row)] = systemData['start']
    worksheet['B{}'.format(row)].font = fontRegularBlack
    worksheet['B{}'.format(row)].alignment = alignmentCenter
    worksheet['C{}'.format(row)] = systemData['10plus-months']
    eval10plusMonths = round(systemData['10plus-months'] / systemData['num-months'], 2)
    worksheet['D{}'.format(row)] = eval10plusMonths

    if eval10plusMonths == 1:
        worksheet['C{}'.format(row)].font = fontRegularBlack
        worksheet['C{}'.format(row)].alignment = alignmentCenter
        worksheet['C{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
        worksheet['D{}'.format(row)].font = fontRegularBlack
        worksheet['D{}'.format(row)].alignment = alignmentCenter
        worksheet['D{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
    elif eval10plusMonths >= 0.6:
        worksheet['C{}'.format(row)].font = fontRegularBlack
        worksheet['C{}'.format(row)].alignment = alignmentCenter
        worksheet['C{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
        worksheet['D{}'.format(row)].font = fontRegularBlack
        worksheet['D{}'.format(row)].alignment = alignmentCenter
        worksheet['D{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
    else:
        worksheet['C{}'.format(row)].font = fontRegularBlack
        worksheet['C{}'.format(row)].alignment = alignmentCenter
        worksheet['D{}'.format(row)].font = fontRegularBlack
        worksheet['D{}'.format(row)].alignment = alignmentCenter

    worksheet['E{}'.format(row)] = systemData['positive-months']
    evalPositiveMonths = round(systemData['positive-months'] / systemData['num-months'], 2)
    worksheet['F{}'.format(row)] = evalPositiveMonths

    if evalPositiveMonths == 1:
        worksheet['E{}'.format(row)].font = fontRegularBlack
        worksheet['E{}'.format(row)].alignment = alignmentCenter
        worksheet['E{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
        worksheet['F{}'.format(row)].font = fontRegularBlack
        worksheet['F{}'.format(row)].alignment = alignmentCenter
        worksheet['F{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
    elif evalPositiveMonths >= 0.7:
        worksheet['E{}'.format(row)].font = fontRegularBlack
        worksheet['E{}'.format(row)].alignment = alignmentCenter
        worksheet['E{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
        worksheet['F{}'.format(row)].font = fontRegularBlack
        worksheet['F{}'.format(row)].alignment = alignmentCenter
        worksheet['F{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
    else:
        worksheet['E{}'.format(row)].font = fontRegularBlack
        worksheet['E{}'.format(row)].alignment = alignmentCenter
        worksheet['F{}'.format(row)].font = fontRegularBlack
        worksheet['F{}'.format(row)].alignment = alignmentCenter

    worksheet['G{}'.format(row)] = systemData['num-months']
    worksheet['G{}'.format(row)].font = fontRegularBlack
    worksheet['G{}'.format(row)].alignment = alignmentCenter
    worksheet['H{}'.format(row)] = systemData['num-picks']
    worksheet['H{}'.format(row)].font = fontRegularBlack
    worksheet['H{}'.format(row)].alignment = alignmentCenter
    monthlyPicks = round(systemData['num-picks'] / systemData['num-months'], 0)
    worksheet['I{}'.format(row)] = monthlyPicks
    worksheet['I{}'.format(row)].font = fontRegularBlack
    worksheet['I{}'.format(row)].alignment = alignmentCenter
    evalPicks = round(monthlyPicks / 20, 2)

    if evalPicks > 1:
        evalPicks = 1
    
    worksheet['J{}'.format(row)] = evalPicks
    worksheet['J{}'.format(row)].font = fontRegularBlack
    worksheet['J{}'.format(row)].alignment = alignmentCenter

    if evalPicks == 1:
        worksheet['I{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
        worksheet['J{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
    elif evalPicks >= 0.9:
        worksheet['I{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
        worksheet['J{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")

    worksheet['K{}'.format(row)] = systemData['yield'] / 100
    worksheet['K{}'.format(row)].font = fontRegularBlack
    worksheet['K{}'.format(row)].alignment = alignmentCenter
    worksheet['K{}'.format(row)].number_format = "0.00%"
    evalYield = round(systemData['yield'] / 20, 2)

    if evalYield > 1:
        evalYield = 1

    worksheet['L{}'.format(row)] = evalYield
    worksheet['L{}'.format(row)].font = fontRegularBlack
    worksheet['L{}'.format(row)].alignment = alignmentCenter

    if evalYield == 1:
        worksheet['K{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
        worksheet['L{}'.format(row)].fill = PatternFill(start_color = "00c0ff", end_color = "00c0ff", fill_type = "solid")
    elif evalYield >= 0.75:
        worksheet['K{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")
        worksheet['L{}'.format(row)].fill = PatternFill(start_color = "c0ff00", end_color = "c0ff00", fill_type = "solid")

    worksheet['M{}'.format(row)] = systemData['start-bank']
    worksheet['M{}'.format(row)].font = fontRegularBlack
    worksheet['M{}'.format(row)].alignment = alignmentCenter
    worksheet['M{}'.format(row)].number_format = "#,##0.00 €".decode('utf-8')
    worksheet['N{}'.format(row)] = systemData['end-bank']
    worksheet['N{}'.format(row)].font = fontRegularBlack
    worksheet['N{}'.format(row)].alignment = alignmentCenter
    worksheet['N{}'.format(row)].number_format = "#,##0.00 €".decode('utf-8')
    profit = systemData['end-bank'] / systemData['start-bank'] - 1
    worksheet['O{}'.format(row)] = profit
    worksheet['O{}'.format(row)].font = fontRegularBlack
    worksheet['O{}'.format(row)].alignment = alignmentCenter
    worksheet['O{}'.format(row)].number_format = "0.00%"

    if profit > 0:
        monthlyProfit = pow(profit * 100, 1 / systemData['num-months']) / 100
    else:
        monthlyProfit = 0

    worksheet['P{}'.format(row)] = monthlyProfit
    worksheet['P{}'.format(row)].font = fontRegularBlack
    worksheet['P{}'.format(row)].alignment = alignmentCenter
    worksheet['P{}'.format(row)].number_format = "0.00%"
    worksheet['Q{}'.format(row)] = pow(monthlyProfit + 1, 12) - 1
    worksheet['Q{}'.format(row)].font = fontRegularBlack
    worksheet['Q{}'.format(row)].alignment = alignmentCenter
    worksheet['Q{}'.format(row)].number_format = "0.00%"
    mark = round(evalPositiveMonths * 4 + evalYield * 3 + eval10plusMonths * 2 + evalPicks, 1)
    worksheet['R{}'.format(row)] = mark
    worksheet['R{}'.format(row)].alignment = alignmentCenter

    if mark >= 9:
        worksheet['R{}'.format(row)].font = fontBoldWhite
        worksheet['R{}'.format(row)].fill = PatternFill(start_color = "5eb91e", end_color = "5eb91e", fill_type = "solid")
    elif mark >= 7:
        worksheet['R{}'.format(row)].font = fontBoldWhite
        worksheet['R{}'.format(row)].fill = PatternFill(start_color = "398ee7", end_color = "c0ff00", fill_type = "solid")
    elif mark >= 5:
        worksheet['R{}'.format(row)].font = fontBoldBlack
        worksheet['R{}'.format(row)].fill = PatternFill(start_color = "ffcc00", end_color = "ffcc00", fill_type = "solid")
    else:
        worksheet['R{}'.format(row)].font = fontBoldWhite
        worksheet['R{}'.format(row)].fill = PatternFill(start_color = "e00000", end_color = "e00000", fill_type = "solid")

    worksheet['T{}'.format(row)] = systemData['global'] / 100
    worksheet['T{}'.format(row)].font = fontRegularBlack
    worksheet['T{}'.format(row)].alignment = alignmentCenter
    worksheet['T{}'.format(row)].number_format = "0.00%"
    avgYield = (systemData['yield'] + systemData['global']) / 200
    worksheet['U{}'.format(row)] = avgYield
    worksheet['U{}'.format(row)].font = fontRegularBlack
    worksheet['U{}'.format(row)].alignment = alignmentCenter
    worksheet['U{}'.format(row)].number_format = "0.00%"
    
    if systemData['num-picks'] >= 250:
        evalPremiuNPicks = 1
    else:
        evalPremiuNPicks = round(systemData['num-picks'] / float(250), 2)
    
    worksheet['V{}'.format(row)] = evalPremiuNPicks
    worksheet['V{}'.format(row)].font = fontRegularBlack
    worksheet['V{}'.format(row)].alignment = alignmentCenter

    if avgYield / 0.1 >= 1:
        evalAvgYield = 1
    else:
        evalAvgYield = round(avgYield / 0.1, 2)
    
    worksheet['W{}'.format(row)] = evalAvgYield
    worksheet['W{}'.format(row)].font = fontRegularBlack
    worksheet['W{}'.format(row)].alignment = alignmentCenter

    if systemData['num-months'] >= 6:
        evalMonths = 1
    else:
        evalMonths = round(float(systemData['num-months']) / 100, 2)
    
    worksheet['X{}'.format(row)] = evalMonths
    worksheet['X{}'.format(row)].font = fontRegularBlack
    worksheet['X{}'.format(row)].alignment = alignmentCenter
    worksheet['Y{}'.format(row)] = evalPicks
    worksheet['Y{}'.format(row)].font = fontRegularBlack
    worksheet['Y{}'.format(row)].alignment = alignmentCenter
    evalPremiuN = evalPremiuNPicks * 2.5 + evalPicks * 2.5 + evalMonths * 2.5 + evalAvgYield * 2.5
    worksheet['Z{}'.format(row)] = evalPremiuN
    worksheet['Z{}'.format(row)].font = fontBoldWhite
    worksheet['Z{}'.format(row)].alignment = alignmentCenter
    worksheet['Z{}'.format(row)].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
    worksheet['Z{}'.format(row)].number_format = "#,##0.000"

    row += 1

worksheet['V{}'.format(row)] = "250p"
worksheet['V{}'.format(row)].font = fontRegularWhite
worksheet['V{}'.format(row)].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
worksheet['V{}'.format(row)].alignment = alignmentCenter
worksheet['W{}'.format(row)] = 0.1
worksheet['W{}'.format(row)].font = fontRegularWhite
worksheet['W{}'.format(row)].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
worksheet['W{}'.format(row)].alignment = alignmentCenter
worksheet['W{}'.format(row)].number_format = "0.00%"
worksheet['X{}'.format(row)] = "6m"
worksheet['X{}'.format(row)].font = fontRegularWhite
worksheet['X{}'.format(row)].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
worksheet['X{}'.format(row)].alignment = alignmentCenter
worksheet['Y{}'.format(row)] = "p/m"
worksheet['Y{}'.format(row)].font = fontRegularWhite
worksheet['Y{}'.format(row)].fill = PatternFill(start_color = "acb20c", end_color = "acb20c", fill_type = "solid")
worksheet['Y{}'.format(row)].alignment = alignmentCenter

'''
    Height: 1 = 0,04 cm.
    Width: 1 = 0,22 cm.
'''
worksheet.row_dimensions[1].height = 25
worksheet.column_dimensions['A'].width = 8.95
worksheet.column_dimensions['B'].width = 10.32
worksheet.column_dimensions['C'].width = 6.82
worksheet.column_dimensions['D'].width = 6.82
worksheet.column_dimensions['E'].width = 7.95
worksheet.column_dimensions['F'].width = 7.95
worksheet.column_dimensions['G'].width = 9.45
worksheet.column_dimensions['H'].width = 5.5
worksheet.column_dimensions['I'].width = 6.82
worksheet.column_dimensions['J'].width = 6.82
worksheet.column_dimensions['K'].width = 7.59
worksheet.column_dimensions['L'].width = 6.82
worksheet.column_dimensions['M'].width = 9.18
worksheet.column_dimensions['N'].width = 8.68
worksheet.column_dimensions['O'].width = 7.73
worksheet.column_dimensions['P'].width = 7.59
worksheet.column_dimensions['Q'].width = 10.45
worksheet.column_dimensions['R'].width = 5.45
worksheet.column_dimensions['S'].width = 1.95
worksheet.column_dimensions['T'].width = 6.86
worksheet.column_dimensions['U'].width = 6.86
worksheet.column_dimensions['V'].width = 7.73
worksheet.column_dimensions['W'].width = 6.86
worksheet.column_dimensions['X'].width = 6.86
worksheet.column_dimensions['Z'].width = 7.73

workbook.save(filepath)