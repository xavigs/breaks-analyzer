# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from os.path import *
from datetime import date, datetime
from models import db, objects

dbConnection = db.Database()
breaksDB = dbConnection.connect()
gamesObj = objects.Games(breaksDB)
dayDateTime = date.today()
dayString = dayDateTime.strftime("%Y-%m-%d")
filepath = "xlsx/{}.xlsx".format(dayString)

if isfile(filepath):
    workbook = openpyxl.load_workbook(filepath)
else:
    workbook = openpyxl.Workbook()

fontRegularBlack = Font(
    name = "Arial",
    size = 10,
    bold = False
)
fontBoldBlack = Font(
    name = "Arial",
    size = 10,
    bold = True
)
fontRegularWhite = Font(
    name = "Arial",
    size = 10,
    bold = False,
    color = "ffffff"
)
alignmentLeft = Alignment(
    horizontal = "left",
    vertical = "center"
)
alignmentCenter = Alignment(
    horizontal = "center",
    vertical = "center"
)
rightBorder = Border(
    right = Side(style="thin")
)

worksheet = workbook.active
worksheet.title = "Daily games"
worksheet['A1'] = dayDateTime
worksheet['A1'].number_format = "dd-MMM"
worksheet['A1'].font = fontBoldBlack
worksheet['A1'].alignment = alignmentCenter
worksheet['A1'].border = rightBorder
worksheet['A1'].fill = PatternFill(start_color = "ffd966", end_color = "ffd966", fill_type = "solid")
games = gamesObj.find_all([{'gameDay': dayString, 'odd': {"$exists": True, "$gte": 1.57}}])

for gameIndex, game in enumerate(games):
    numRow = gameIndex + 1
    worksheet['B{}'.format(numRow)] = "ATP"
    worksheet['B{}'.format(numRow)].font = fontRegularBlack
    worksheet['B{}'.format(numRow)].alignment = alignmentCenter
    worksheet['B{}'.format(numRow)].fill = PatternFill(start_color = "ccccdd", end_color = "ccccdd", fill_type = "solid")
    worksheet['C{}'.format(numRow)] = 500
    worksheet['C{}'.format(numRow)].font = fontRegularBlack
    worksheet['C{}'.format(numRow)].alignment = alignmentCenter
    worksheet['C{}'.format(numRow)].fill = PatternFill(start_color = "ccccdd", end_color = "ccccdd", fill_type = "solid")
    worksheet['D{}'.format(numRow)] = "I"
    worksheet['D{}'.format(numRow)].font = fontRegularWhite
    worksheet['D{}'.format(numRow)].alignment = alignmentCenter
    worksheet['D{}'.format(numRow)].fill = PatternFill(start_color = "618cb1", end_color = "618cb1", fill_type = "solid")
    worksheet['E{}'.format(numRow)] = "Text1"
    worksheet['E{}'.format(numRow)].font = fontBoldBlack
    worksheet['E{}'.format(numRow)].alignment = alignmentLeft
    worksheet['F{}'.format(numRow)] = "Text2"
    worksheet['F{}'.format(numRow)].font = fontRegularBlack
    worksheet['F{}'.format(numRow)].alignment = alignmentLeft
    worksheet['G{}'.format(numRow)] = "OK"
    worksheet['G{}'.format(numRow)].font = fontRegularWhite
    worksheet['G{}'.format(numRow)].alignment = alignmentCenter
    worksheet['G{}'.format(numRow)].fill = PatternFill(start_color = "34a853", end_color = "34a853", fill_type = "solid")
    worksheet['H{}'.format(numRow)] = 1.61
    worksheet['H{}'.format(numRow)].font = fontRegularBlack
    worksheet['H{}'.format(numRow)].alignment = alignmentCenter
    worksheet['I{}'.format(numRow)] = "N"
    worksheet['I{}'.format(numRow)].font = fontRegularWhite
    worksheet['I{}'.format(numRow)].alignment = alignmentCenter
    worksheet['I{}'.format(numRow)].fill = PatternFill(start_color = "ee0000", end_color = "ee0000", fill_type = "solid")
    worksheet['J{}'.format(numRow)] = "S"
    worksheet['J{}'.format(numRow)].font = fontRegularWhite
    worksheet['J{}'.format(numRow)].alignment = alignmentCenter
    worksheet['J{}'.format(numRow)].fill = PatternFill(start_color = "ee0000", end_color = "ee0000", fill_type = "solid")
    worksheet['K{}'.format(numRow)] = 1.66
    worksheet['K{}'.format(numRow)].font = fontRegularBlack
    worksheet['K{}'.format(numRow)].alignment = alignmentCenter
    worksheet['L{}'.format(numRow)].value = "74,50%"
    worksheet['L{}'.format(numRow)].number_format = "0.00%"
    worksheet['L{}'.format(numRow)].font = fontBoldBlack
    worksheet['L{}'.format(numRow)].alignment = alignmentCenter
    worksheet['M{}'.format(numRow)].value = "=1/K1" # 62,11%
    worksheet['M{}'.format(numRow)].number_format = "0.00%"
    worksheet['M{}'.format(numRow)].font = fontRegularBlack
    worksheet['M{}'.format(numRow)].alignment = alignmentCenter
    worksheet['N{}'.format(numRow)] = "=L1/M1"
    worksheet['N{}'.format(numRow)].font = fontBoldBlack
    worksheet['N{}'.format(numRow)].alignment = alignmentCenter
    worksheet['N{}'.format(numRow)].number_format = "#,##0.00"
    worksheet['B{}'.format(numRow)].border = rightBorder
    
    '''worksheet['B2'] = "CH"
    worksheet['B2'].font = fontRegularWhite
    worksheet['B2'].alignment = alignmentCenter
    worksheet['B2'].fill = PatternFill(start_color = "53a01d", end_color = "53a01d", fill_type = "solid")
    worksheet.merge_cells('B2:C2')
    worksheet['D2'] = "D"
    worksheet['D2'].font = fontRegularWhite
    worksheet['D2'].alignment = alignmentCenter
    worksheet['D2'].fill = PatternFill(start_color = "788575", end_color = "788575", fill_type = "solid")
    worksheet['E2'] = "Text3"
    worksheet['E2'].font = fontBoldBlack
    worksheet['E2'].alignment = alignmentLeft
    worksheet['F2'] = "Text4"
    worksheet['F2'].font = fontRegularBlack
    worksheet['F2'].alignment = alignmentLeft
    worksheet['G2'] = "10"
    worksheet['G2'].font = fontRegularBlack
    worksheet['G2'].alignment = alignmentCenter
    worksheet['H2'] = 1.61
    worksheet['H2'].font = fontRegularBlack
    worksheet['H2'].alignment = alignmentCenter
    worksheet['I2'] = "S"
    worksheet['I2'].font = fontRegularWhite
    worksheet['I2'].alignment = alignmentCenter
    worksheet['I2'].fill = PatternFill(start_color = "34a853", end_color = "34a853", fill_type = "solid")
    worksheet['J2'] = "N"
    worksheet['J2'].font = fontRegularWhite
    worksheet['J2'].alignment = alignmentCenter
    worksheet['J2'].fill = PatternFill(start_color = "34a853", end_color = "34a853", fill_type = "solid")
    worksheet['K2'] = 1.66
    worksheet['K2'].font = fontRegularBlack
    worksheet['K2'].alignment = alignmentCenter
    worksheet['L2'].value = "74,50%"
    worksheet['L2'].number_format = "0.00%"
    worksheet['L2'].font = fontBoldBlack
    worksheet['L2'].alignment = alignmentCenter
    worksheet['M2'].value = "=1/K1" # 62,11%
    worksheet['M2'].number_format = "0.00%"
    worksheet['M2'].font = fontRegularBlack
    worksheet['M2'].alignment = alignmentCenter
    worksheet['N2'] = "=L1/M1"
    worksheet['N2'].font = fontBoldBlack
    worksheet['N2'].alignment = alignmentCenter
    worksheet['N2'].number_format = "#,##0.00"'''

worksheet.merge_cells('A1:A2')
workbook.save(filepath)

'''
    # TODO DB

    1) Collection tournament (teID)
    2) Save tournament data: teID (_id), teName, sex, category, surface
    3) Into game documents, new fields: tournament (teID), time
    4) Into player lastGames (or game documents), newField: last ("OK", "01", "02", "10", "11", "12", "20", "21", "22")
    5) Tractament d'excepcions a tots els scripts, perquè es generi un log, però no deixi de funcionar quan un element falla
'''
